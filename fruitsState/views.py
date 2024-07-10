from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden, HttpResponse
from .forms import CreateUser
from django.utils.translation import activate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.views.generic.edit import CreateView, UpdateView
from django.views.generic import TemplateView
from .models import ImageModel, FruitCount
from .forms import ImageUploadForm, ImageEditForm
from urllib.error import HTTPError
import time
import os
from django.core.files import File
from django.conf import settings
import torch
import pathlib
import numpy
from PIL import Image
from openpyxl import Workbook

# Create your views here.


class SingletonModel:
    _instance = None

    @staticmethod
    def get_instance():
        if SingletonModel._instance is None:
            SingletonModel()
        return SingletonModel._instance

    def __init__(self):
        if SingletonModel._instance is not None:
            raise Exception("This class is a singleton!")
        else:
            # Deben cambiar esta ruta por la donde esta su red neuronal
            folder_path = 'C:/Users/jairg/Desktop/Frutas IA/fruitsState/IA/IAfinal.pt'
            self.model = torch.hub.load('ultralytics/yolov5', 'custom', path=folder_path, force_reload=True)
            SingletonModel._instance = self



def Register(request):
    activate('es')
    if request.method == 'POST':
        form = CreateUser(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = CreateUser()

    return render(request, 'register.html', {'form': form})



def Home(request):
    user = request.user

    if not user.is_authenticated:
        return redirect('login')
    
    images = ImageModel.objects.all()

    stats = {
        'total_good': 0,
        'total_bad': 0,        
        'total': 0,
        'apple': {'good': 0, 'bad': 0},
        'banana': {'good': 0, 'bad': 0},
        'orange': {'good': 0, 'bad': 0},
        'pomegranate': {'good': 0, 'bad': 0},
    }

    # Calcular estadísticas
    for image in images:
        stats['total_good'] += image.good_apple + image.good_banana + image.good_orange + image.good_pomegranate
        stats['total_bad'] += image.bad_apple + image.bad_banana + image.bad_orange + image.bad_pomegranate

        stats['total'] += image.good_apple + image.good_banana + image.good_orange + image.good_pomegranate + image.bad_apple + image.bad_banana + image.bad_orange + image.bad_pomegranate
        stats['apple']['good'] += image.good_apple
        stats['apple']['bad'] += image.bad_apple

        stats['banana']['good'] += image.good_banana
        stats['banana']['bad'] += image.bad_banana

        stats['orange']['good'] += image.good_orange
        stats['orange']['bad'] += image.bad_orange

        stats['pomegranate']['good'] += image.good_pomegranate
        stats['pomegranate']['bad'] += image.bad_pomegranate
    return render(request,'home.html', {
        'user': user,
        'stats': stats
    })

class GenerateExcell(TemplateView):
    def get(self, request, *args, **kwargs):
        dataAnalized = ImageModel.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de datos analizados'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'Frutas Analizadas'
        ws['C3'] = 'Aptos para vender'
        ws['D3'] = 'No apto para venta'
        ws['E3'] = 'Naranjas'
        ws['F3'] = 'Granadas'
        ws['G3'] = 'Bananas'
        ws['H3'] = 'Manzanas'
        ws['I3'] = 'Naranjas Malas'
        ws['J3'] = 'Naranjas Buenas'
        ws['K3'] = 'Granadas Malas'
        ws['L3'] = 'Granadas Buenas'
        ws['M3'] = 'Bananas Malas'
        ws['N3'] = 'Bananas Buenas'
        ws['O3'] = 'Manzanas Malas'
        ws['P3'] = 'Manzanas Buenas'

        stats = {
        'total_good': 0,
        'total_bad': 0,        
        'total': 0,
        'apple': {'good': 0, 'bad': 0},
        'banana': {'good': 0, 'bad': 0},
        'orange': {'good': 0, 'bad': 0},
        'pomegranate': {'good': 0, 'bad': 0},
        }


        for data in dataAnalized:
            stats['total_good'] += data.good_apple + data.good_banana + data.good_orange + data.good_pomegranate
            stats['total_bad'] += data.bad_apple + data.bad_banana + data.bad_orange + data.bad_pomegranate

            stats['total'] += data.good_apple + data.good_banana + data.good_orange + data.good_pomegranate + data.bad_apple + data.bad_banana + data.bad_orange + data.bad_pomegranate
            stats['apple']['good'] += data.good_apple
            stats['apple']['bad'] += data.bad_apple

            stats['banana']['good'] += data.good_banana
            stats['banana']['bad'] += data.bad_banana

            stats['orange']['good'] += data.good_orange
            stats['orange']['bad'] += data.bad_orange

            stats['pomegranate']['good'] += data.good_pomegranate
            stats['pomegranate']['bad'] += data.bad_pomegranate
        


       
        ws.cell(row=4, column=2).value = stats['total']
        ws.cell(row=4, column=3).value = stats['total_good']
        ws.cell(row=4, column=4).value = stats['total_bad']
        ws.cell(row=4, column=5).value = stats['orange']['good'] + stats['orange']['bad']
        ws.cell(row=4, column=6).value = stats['pomegranate']['good'] + stats['pomegranate']['bad']
        ws.cell(row=4, column=7).value = stats['banana']['good'] + stats['banana']['bad']
        ws.cell(row=4, column=8).value = stats['apple']['good'] + stats['apple']['bad']
        ws.cell(row=4, column=9).value = stats['orange']['bad']
        ws.cell(row=4, column=10).value = stats['orange']['good']
        ws.cell(row=4, column=11).value = stats['pomegranate']['bad']
        ws.cell(row=4, column=12).value = stats['pomegranate']['good']
        ws.cell(row=4, column=13).value = stats['banana']['bad']
        ws.cell(row=4, column=14).value = stats['banana']['good']
        ws.cell(row=4, column=15).value = stats['apple']['bad']
        ws.cell(row=4, column=16).value = stats['apple']['good']

        archive_name = 'Reporte.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename={0}".format(archive_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class GenerateExcellPerUser(TemplateView):
    def get(self, request, *args, **kwargs):
        dataAnalized = ImageModel.objects.filter(user_id=request.user.id)
        wb = Workbook()
        ws = wb.active
        ws['B1'] = f'Reporte analizados de usuario {request.user}'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'Frutas Analizadas'
        ws['C3'] = 'Aptos para vender'
        ws['D3'] = 'No apto para venta'
        ws['E3'] = 'Naranjas'
        ws['F3'] = 'Granadas'
        ws['G3'] = 'Bananas'
        ws['H3'] = 'Manzanas'
        ws['I3'] = 'Naranjas Malas'
        ws['J3'] = 'Naranjas Buenas'
        ws['K3'] = 'Granadas Malas'
        ws['L3'] = 'Granadas Buenas'
        ws['M3'] = 'Bananas Malas'
        ws['N3'] = 'Bananas Buenas'
        ws['O3'] = 'Manzanas Malas'
        ws['P3'] = 'Manzanas Buenas'

        stats = {
        'total_good': 0,
        'total_bad': 0,        
        'total': 0,
        'apple': {'good': 0, 'bad': 0},
        'banana': {'good': 0, 'bad': 0},
        'orange': {'good': 0, 'bad': 0},
        'pomegranate': {'good': 0, 'bad': 0},
        }


        for data in dataAnalized:
            stats['total_good'] += data.good_apple + data.good_banana + data.good_orange + data.good_pomegranate
            stats['total_bad'] += data.bad_apple + data.bad_banana + data.bad_orange + data.bad_pomegranate

            stats['total'] += data.good_apple + data.good_banana + data.good_orange + data.good_pomegranate + data.bad_apple + data.bad_banana + data.bad_orange + data.bad_pomegranate
            stats['apple']['good'] += data.good_apple
            stats['apple']['bad'] += data.bad_apple

            stats['banana']['good'] += data.good_banana
            stats['banana']['bad'] += data.bad_banana

            stats['orange']['good'] += data.good_orange
            stats['orange']['bad'] += data.bad_orange

            stats['pomegranate']['good'] += data.good_pomegranate
            stats['pomegranate']['bad'] += data.bad_pomegranate
        


       
        ws.cell(row=4, column=2).value = stats['total']
        ws.cell(row=4, column=3).value = stats['total_good']
        ws.cell(row=4, column=4).value = stats['total_bad']
        ws.cell(row=4, column=5).value = stats['orange']['good'] + stats['orange']['bad']
        ws.cell(row=4, column=6).value = stats['pomegranate']['good'] + stats['pomegranate']['bad']
        ws.cell(row=4, column=7).value = stats['banana']['good'] + stats['banana']['bad']
        ws.cell(row=4, column=8).value = stats['apple']['good'] + stats['apple']['bad']
        ws.cell(row=4, column=9).value = stats['orange']['bad']
        ws.cell(row=4, column=10).value = stats['orange']['good']
        ws.cell(row=4, column=11).value = stats['pomegranate']['bad']
        ws.cell(row=4, column=12).value = stats['pomegranate']['good']
        ws.cell(row=4, column=13).value = stats['banana']['bad']
        ws.cell(row=4, column=14).value = stats['banana']['good']
        ws.cell(row=4, column=15).value = stats['apple']['bad']
        ws.cell(row=4, column=16).value = stats['apple']['good']

        archive_name = 'Reporte.xlsx'
        response = HttpResponse(content_type='application/ms-excel')
        content = "attachment; filename={0}".format(archive_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


def Login(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username= username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Bienvenido {username}')
                return redirect('home')
            else:
                messages.error(request, 'Usuario o contraseña incorrectos')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')


    form = AuthenticationForm()   

    form.fields['username'].label = 'Nombre de usuario'
    form.fields['password'].label = 'Contraseña'
    return render(request, 'login.html', {"form":form})



def Logout(request):
    logout(request)
    return redirect('login')


class DetectFruits(CreateView):
    model = ImageModel
    template_name = 'detectFruit.html'
    fields = ["image"]

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            form = ImageUploadForm(request.POST, request.FILES)
            if form.is_valid():
                fruitImage = form.save(commit=False)
                fruitImage.user_id = request.user

                try:
                    img_path = fruitImage.image
                    imageByIa, bad_apple, bad_banana, bad_orange, bad_pomegranate, good_apple, good_banana, good_orange, good_pomegranate= fruitState(img_path)
                except HTTPError as e:
                    if e.code == 403:
                        print("Rate limit exceeded. Waiting before retrying...")
                        time.sleep(200)  # Esperar 60 segundos antes de reintentar
                        return fruitState(fruitImage.image.path)
                    else:
                        raise

                print(bad_pomegranate)
                fruitImage.bad_apple = bad_apple
                fruitImage.bad_banana = bad_banana
                fruitImage.bad_orange = bad_orange
                fruitImage.bad_pomegranate = bad_pomegranate
                fruitImage.good_apple = good_apple
                fruitImage.good_banana = good_banana
                fruitImage.good_orange = good_orange
                fruitImage.good_pomegranate = good_pomegranate

                fruitsCount = FruitCount.objects.create()


                if bad_apple >= 1:
                    fruitImage.fruitsState += f'Manzana M:{bad_apple}\n,'
                    fruitsCount.appleBad += bad_apple

                if bad_banana >= 1:
                    fruitImage.fruitsState += f'Platana M:{bad_banana} \n,'
                    fruitsCount.bananaBad += bad_banana

                if bad_orange >= 1:
                    fruitImage.fruitsState += f'Naranja M:{bad_orange} \n,'
                    fruitsCount.orangeBad += bad_orange

                if bad_pomegranate >= 1:
                    fruitImage.fruitsState += f'Granada M:{bad_pomegranate} \n,'
                    fruitsCount.pomegranateBad += bad_pomegranate

                if good_apple >= 1:
                    fruitImage.fruitsState += f'Manzana B:{good_apple} \n,'
                    fruitsCount.appleGood += good_apple

                if good_banana >= 1:
                    fruitImage.fruitsState += f'Platano B:{good_banana} \n,'
                    fruitsCount.bananaGood += good_banana

                if good_orange >= 1:
                    fruitImage.fruitsState += f'Naranja B:{good_orange} \n,'
                    fruitsCount.orangeGood += good_orange

                if good_pomegranate >= 1:
                    fruitImage.fruitsState += f'Granada B:{good_pomegranate} \n,,'
                    fruitsCount.pomegranateGood += good_pomegranate

                fruitsCount.save()
                fruitImage.fruitCount_id = fruitsCount
                processed_image_dir = settings.DETECTION_MEDIA_ROOT
                if not os.path.exists(processed_image_dir):
                    os.makedirs(processed_image_dir)

                processed_image_path = os.path.join(processed_image_dir, 'deteccion.png')
                imageByIa.save(processed_image_path)
                processed_image = File(open(processed_image_path, 'rb'))
                fruitImage.imagesDetected.save('deteccion.png', processed_image, save=False)

                fruitImage.image.name = 'mydetect.png'
                fruitImage.save()

                user_now = request.user.id
                return redirect('inventory', user_id=user_now)
        else:
            form = ImageUploadForm()
        return render(request, 'detectFruit.html', {'form': form})


class EditDetectFruits(UpdateView):
    model = ImageModel
    template_name = 'editDetectFruit.html'
    form_class = ImageEditForm

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.form_class(request.POST, request.FILES, instance=self.object)
        if form.is_valid():
            fruitImage = form.save(commit=False)
            fruitImage.user_id = request.user
            fruitsCount = fruitImage.fruitCount_id 

            fruitsCount.bananaBad= fruitImage.bad_banana
            fruitsCount.appleBad= fruitImage.bad_apple
            fruitsCount.orangeBad= fruitImage.bad_orange
            fruitsCount.pomegranateBad = fruitImage.bad_pomegranate
            fruitsCount.appleGood= fruitImage.good_apple
            fruitsCount.bananaGood= fruitImage.good_banana
            fruitsCount.orangeGood= fruitImage.good_orange
            fruitsCount.pomegranateGood= fruitImage.good_pomegranate

            fruitsCount.save()
            fruitImage.save()
            return redirect('inventory', user_id=request.user.id)
        return render(request, self.template_name, {
            'form': form,
        })
    
    def get_object(self, queryset=None):
        return get_object_or_404(ImageModel, pk=self.kwargs.get('pk'))



def DeleteRegisterFruit(request, user_id):
    imgToDelete = get_object_or_404(ImageModel, pk=user_id)
    if request.user == imgToDelete.user_id:
        fruitsCount = imgToDelete.fruitCount_id
        if fruitsCount:
            fruitsCount.bananaBad -= imgToDelete.bad_banana
            fruitsCount.appleBad -= imgToDelete.bad_apple
            fruitsCount.orangeBad -= imgToDelete.bad_orange
            fruitsCount.pomegranateBad  -= imgToDelete.bad_pomegranate
            fruitsCount.appleGood -= imgToDelete.good_apple
            fruitsCount.bananaGood -= imgToDelete.good_banana
            fruitsCount.orangeGood -= imgToDelete.good_orange
            fruitsCount.pomegranateGood -= imgToDelete.good_pomegranate
        imgToDelete.delete()
        return redirect('inventory', user_id=request.user.id)
    else:
        return HttpResponseForbidden('No tienes permiso de editar estos datos')







def fruitState(img_path):
    temp = pathlib.PosixPath   
    pathlib.PosixPath = pathlib.WindowsPath
    model_instance = SingletonModel.get_instance()
    model = model_instance.model
    img = Image.open(img_path)
    results = model(img)
    results.show()
    r_img = results.render()
    img_with_boxes = r_img[0]
    new_img = Image.fromarray(img_with_boxes) 

    fruits = []
    for fruta in results.xyxy[0]:
        x0, y0, x1, y1, confi, cla = fruta.numpy()
        fruits.append({
                'x0': x0,
                'y0': y0,
                'x1': x1,
                'y1': y1,
                'confianza': confi,
                'clase': cla
            })

    numBadApple = 0
    numBadOrange = 0
    numBadBanana = 0
    numBadPomegranate = 0
    numGoodApple = 0
    numGoodOrange = 0
    numGoodBanana = 0
    numGoodPomegranate = 0
    
    for dataFruits in fruits:
        # define el nivel de confianza para filtrar solo las detecciones q cumplen con ese dato
        if dataFruits['confianza'] > 0.60:
            if dataFruits['clase'] == 0:
                numBadApple += 1

            if dataFruits['clase'] == 1:
                numBadBanana += 1

            if dataFruits['clase'] == 2:
                numBadOrange +=1

            if dataFruits['clase'] == 3:
                numBadPomegranate += 1
            
            if dataFruits['clase'] == 4:
                numGoodApple += 1
            
            if dataFruits['clase'] == 5:
                numGoodBanana += 1

            if dataFruits['clase'] == 6:
                numGoodOrange += 1

            if dataFruits['clase'] == 7:
                numGoodPomegranate += 1
        
    
    # print(fruits)

    return new_img, numBadApple, numBadBanana, numBadOrange, numBadPomegranate, numGoodApple, numGoodBanana, numGoodOrange, numGoodPomegranate







def Inventory(request, user_id):
    user = request.user
    inventory = ImageModel.objects.filter(user_id=user_id)
    print(inventory)
    return render(request, 'inventory.html', {
        'user': user,
        'inventory': inventory
    })


def Clasification(request, user_id):
    user_id = request.user
    dataDetected = ImageModel.objects.filter(user_id=user_id)
    
    total_bad_apple = 0
    total_bad_banana = 0
    total_bad_orange = 0
    total_bad_pomegranate = 0

    total_good_apple = 0
    total_good_banana = 0
    total_good_orange = 0
    total_good_pomegranate = 0

    for data in dataDetected:
        total_bad_apple += data.bad_apple
        total_bad_banana += data.bad_banana
        total_bad_orange += data.bad_orange
        total_bad_pomegranate += data.bad_pomegranate

        total_good_apple += data.good_apple
        total_good_banana += data.good_banana
        total_good_orange += data.good_orange
        total_good_pomegranate += data.good_pomegranate

    total_apple = total_bad_apple + total_good_apple
    total_banana = total_good_banana + total_bad_banana
    total_orange = total_good_orange + total_bad_orange
    total_pomegranate = total_bad_pomegranate + total_good_pomegranate

    context = {
        'total_bad_apple': total_bad_apple,
        'total_bad_banana': total_bad_banana,
        'total_bad_orange': total_bad_orange,
        'total_bad_pomegranate': total_bad_pomegranate,
        'total_good_apple': total_good_apple,
        'total_good_banana': total_good_banana,
        'total_good_orange': total_good_orange,
        'total_good_pomegranate': total_good_pomegranate,
        'total_apple': total_apple,
        'total_banana': total_banana,
        'total_orange': total_orange,
        'total_pomegranate': total_pomegranate,
    }

    return render(request, 'clasification.html', context)